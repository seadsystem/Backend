#!/usr/bin/python
# ============================================================
# File: harmonics_3d.py
# Description: Takes a xlsx file from the PA1000 and 3d plots
#              the Voltage, Current, and Wattage harmonics
#              usage: harmonics_3d.py [source.xlsx]
# Created by Henry Crute
# 9/12/2014
# ============================================================

from mpl_toolkits.mplot3d import Axes3D

import os
import datetime
import sys
import random
import math
import numpy as np
import matplotlib.pyplot as plt

from openpyxl import load_workbook

#creates folder if it doesn't exist in the directory
def ensure_dir(directory):
   if not os.path.exists(directory):
      os.makedirs(directory)

#uses openpyxl library to parse all of the data in voltage, amperage
#input variables. extremely ugly
def get_harmonics(filename, voltage, amperage, wattage):
   wb2 = load_workbook(filename,  use_iterators = True)
   ws = wb2.get_sheet_by_name(name = 'PA1000 (0122) Ch1')
   voltIndex = []
   ampIndex = []
   j = -2
   #iterate through all rows
   for row in ws.iter_rows():
      i = -1
      if j > -2:
         amperage.append([])
         voltage.append([])
      j = j + 1
      #iterate through each cell in row
      for cell in row:
         i = i + 1
         #extract data from the index values and puts it into arrays
         if i in voltIndex:
            if cell.value == '':
               continue
            #print cell.value
            voltage[j].append(cell.value)
            continue
         if i in ampIndex:
            if cell.value == '':
               continue
            #print cell.value
            amperage[j].append(cell.value)
            continue
         if j > 0 and i == 3:
            wattage.append(cell.value)
         #placeholder index values for the first row, tells where data is
         if type(cell.value) is datetime.datetime:
            continue
         elif type(cell.value) is float:
            continue
         elif 'Volt' in cell.value and 'Magnitude' in cell.value:
            #print cell.value
            voltIndex.append(i)
         elif 'Magnitude' in cell.value and 'Amp' in cell.value:
            #print cell.value
            ampIndex.append(i)

#takes average of given arrays, from start_index to stop_index
#also finds max, and min values within the range
def get_avgmaxmin(input_list, start_index, stop_index, mean_list, std_list):
   if stop_index == 0:
      return
   #initialize array of zeros size equal to number of harmonics
   mean = [0] * len(input_list[0])
   for index in range(start_index, stop_index):
      for number in range(len(input_list[index])):
         mean[number] = mean[number] + input_list[index][number]
   for number in range(len(mean)):
      mean[number] = mean[number] / (stop_index - start_index)
   #print mean
   mean_list.append(mean)
   #initialize array of zeros equal to number of harmonics for variance
   variance = [0] * len(input_list[0])
   for number in range(len(input_list[0])):
      for index in range(start_index, stop_index):
         variance[number] = variance[number] + math.pow((input_list[index][number] - mean[number]), 2)
   for number in range(len(variance)):
      variance[number] = variance[number] / (stop_index - start_index)
   #print variance
   stdeviation = [0] * len(input_list[0])
   for number in range(len(variance)):
      stdeviation[number] = math.sqrt(variance[number])
   std_list.append(stdeviation)
   #print stdeviation

#percent difference comparison to see transients in snapshot, and trim
def trim_harmonics(voltage, amperage, volt_mean, amp_mean, volt_std, amp_std):
   #for voltage
   i = 0
   prevSum = -1
   averageStart = 0
   averageStop = 0
   #iterates over a copy of the list
   for row in list(voltage):
      curSum = sum(row)
      #solves division by zero error
      if ((curSum + prevSum) == 0):
         per_diff = 0
      else:
         per_diff = abs(curSum - prevSum)/((curSum + prevSum)/2)
      #print 'per_diff = ' + str(per_diff)
      if abs(per_diff) < 0.10: #########################CHANGEME
         #print 'deleting ' + str(voltage[i])
         #voltage.remove(row)
         pass
      else :
         get_avgmaxmin(voltage, averageStart, averageStop, volt_mean, volt_std)
         averageStart = averageStop
      prevSum = curSum
      i = i + 1
      averageStop = averageStop + 1
   get_avgmaxmin(amperage, averageStart, averageStop, volt_mean, volt_std)

   #for amperage
   i = 0
   prevSum = -1
   averageStart = 0
   averageStop = 0
   #iterates over a copy of the list
   for row in list(amperage):
      curSum = sum(row)
      #solves division by zero error
      if ((curSum + prevSum) == 0):
         per_diff = 0
      else:
         per_diff = abs(curSum - prevSum)/((curSum + prevSum)/2)
      #print 'per_diff = ' + str(per_diff)
      if abs(per_diff) < 0.10: #########################CHANGEME
         #print 'deleting row ' + str(averageStop)
         #amperage.remove(row)
         pass
      else :
         get_avgmaxmin(amperage, averageStart, averageStop, amp_mean, amp_std)
         averageStart = averageStop
      prevSum = curSum
      i = i + 1
      averageStop = averageStop + 1
   get_avgmaxmin(amperage, averageStart, averageStop, amp_mean, amp_std)

def visualize_data(wattage, filename):
   ensure_dir(output_dir + filename)
   fig = plt.figure()
   plt.title('Power from ' + filename)
   plt.xlabel('Seconds')
   plt.ylabel('Watts')
   x = np.arange(0, len(wattage) / 2.0, 0.5)
   plt.plot(x, wattage, 'r',
            x, wattage, 'ro')
   fig.savefig(output_dir + filename + '/' + filename + '_watt' + '.png')

#creates a 3 dimensional bargraph from two dimensional list_in
#with colors specified from input colorpicker
def bargraph_3d(list_in, list_std, colorpicker, filename):
   ensure_dir(output_dir + filename)
   fig = plt.figure()
   ax = fig.add_subplot(111, projection='3d')
   ax.set_title('Harmonics from ' + filename)
   yaxis = np.arange(len(list_in))
   #print yaxis
   print "colorpicker is " + str(len(colorpicker))
   print "yaxis is " + str(len(yaxis))
   for c, z in zip(colorpicker, yaxis):
      xs = np.arange(len(list_in[z]))
      #print xs
      ys = list_in[z]
      #print ys
      cs = c
      #make colors different, e.g. each
      ax.bar(xs, ys, zs=z, zdir='y', color=c, alpha=0.8)
   ax.set_zlim(0,1)
   ax.set_xlabel('Harmonic Number')
   ax.set_ylabel('Change over Time')
   ax.set_zlabel('Harmonic Arms Amplitude')
   fig.savefig(output_dir + filename + '/' + filename + '_harmonics.png')
   plt.show()

def process_data(file_directory):
   filename = os.path.basename(file_directory)
   #2 dimensional list initializations
   voltage = []
   amperage = []
   wattage = []
   #grabs all of the harmonics from the xlsx
   get_harmonics(file_directory, voltage, amperage, wattage)
   #trims harmonics to get means, and standard deviations to graph
   volt_mean = []
   amp_mean = []
   watt_mean = []
   volt_std = []
   amp_std = []
   watt_std = []
   trim_harmonics(voltage, amperage, volt_mean, amp_mean, volt_std, amp_std)
   print amp_mean
   #creates a random color list for voltage then amperage
   volt_color = []
   for x in range(0, len(volt_mean)):
      volt_color.append((random.random(), random.random(), random.random()))
   amp_color = []
   for x in range(0, len(amp_mean)):
      amp_color.append((random.random(), random.random(), random.random()))
   #make bar graphs
   visualize_data(wattage, filename)
   #bargraph_3d(volt_mean, amp_std, volt_color)
   bargraph_3d(amp_mean, amp_std, amp_color, filename)

#START OF PROGRAM
if len(sys.argv) < 3:
   print 'usage: ' + sys.argv[0] + ' [input directory] [output directory]'
   exit(1)
#creates workbook from input argument
path = sys.argv[1]

#opens output directory to output files
output_dir = sys.argv[2]

for root, dirs, filenames in os.walk(path):
   for f in filenames:
      absolute_path = path + f
      print absolute_path
      process_data(absolute_path)

